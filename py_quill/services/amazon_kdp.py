"""KDP xlsx parsing helpers for daily sales and KENP stats."""

from __future__ import annotations

import datetime
import io
from typing import Any

import openpyxl
from common import book_defs, models

_CURRENCY_CODE_TO_USD_RATE: dict[str, float] = {
  "USD": 1.0,
  "CAD": 0.7320,
  "GBP": 1.3569,
}
_MARKETPLACE_TO_MARKET_CODE: dict[str, str] = {
  "amazon.com": "US",
  "amazon.ca": "CA",
  "amazon.co.uk": "GB",
  "amazon.com.au": "AU",
  "amazon.de": "DE",
  "amazon.fr": "FR",
  "amazon.es": "ES",
  "amazon.it": "IT",
  "amazon.co.jp": "JP",
  "amazon.com.mx": "MX",
  "amazon.nl": "NL",
  "amazon.se": "SE",
  "amazon.pl": "PL",
  "amazon.in": "IN",
  "amazon.com.br": "BR",
}
_MARKET_CODE_TO_CURRENCY_CODE: dict[str, str] = {
  "US": "USD",
  "CA": "CAD",
  "GB": "GBP",
  "UK": "GBP",
}


class AmazonKdpError(Exception):
  """Raised when KDP report parsing or validation fails."""


def parse_kdp_xlsx(file_bytes: bytes) -> list[models.AmazonKdpDailyStats]:
  """Parse a KDP dashboard xlsx into daily stats models."""
  if not file_bytes:
    raise AmazonKdpError("KDP xlsx upload was empty")

  workbook = _load_workbook(file_bytes)
  combined_sales = _read_sheet_rows(workbook, "Combined Sales")
  kenp_rows = _read_sheet_rows(workbook, "KENP Read")

  stats_by_date: dict[datetime.date, models.AmazonKdpDailyStats] = {}
  _apply_combined_sales_rows(combined_sales, stats_by_date)
  _apply_kenp_rows(kenp_rows, stats_by_date)
  _finalize_stats(stats_by_date)
  return [stats_by_date[date_value] for date_value in sorted(stats_by_date)]


def _load_workbook(file_bytes: bytes) -> openpyxl.Workbook:
  """Load an xlsx workbook from in-memory bytes."""
  try:
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
  except Exception as exc:
    raise AmazonKdpError(f"Failed to parse xlsx: {exc}") from exc


def _read_sheet_rows(workbook: openpyxl.Workbook,
                     sheet_name: str) -> list[dict[str, Any]]:
  """Read a sheet into a list of dict rows keyed by header."""
  if sheet_name not in workbook.sheetnames:
    raise AmazonKdpError(f"Missing required sheet: {sheet_name}")
  sheet = workbook[sheet_name]
  all_rows = list(sheet.iter_rows(values_only=True))
  if not all_rows:
    return []

  headers_raw = all_rows[0]
  headers = [str(h).strip() if h is not None else "" for h in headers_raw]
  rows: list[dict[str, Any]] = []
  for row in all_rows[1:]:
    if all(cell is None or str(cell).strip() == "" for cell in row):
      continue
    row_dict: dict[str, Any] = {}
    for index, header in enumerate(headers):
      if not header:
        continue
      row_dict[header] = row[index] if index < len(row) else None
    rows.append(row_dict)
  return rows


def _apply_combined_sales_rows(
  rows: list[dict[str, Any]],
  stats_by_date: dict[datetime.date, models.AmazonKdpDailyStats],
) -> None:
  """Merge Combined Sales rows into daily aggregates with validations."""
  for row in rows:
    date_value = _parse_iso_date(row.get("Royalty Date"), "Royalty Date")
    asin_or_isbn = _required_str(row.get("ASIN/ISBN"), "ASIN/ISBN")
    market = _parse_market(row.get("Marketplace"))
    currency_code = _required_str(row.get("Currency"), "Currency").upper()

    book_variant = book_defs.find_book_variant(asin_or_isbn)
    if book_variant is None:
      raise AmazonKdpError(f"Unknown ASIN/ISBN in KDP report: {asin_or_isbn}")

    format_name = _format_from_transaction_type(row.get("Transaction Type"))
    if book_variant.format.value != format_name:
      raise AmazonKdpError(
        f"Format mismatch for {asin_or_isbn}: report={format_name} book_defs={book_variant.format.value}"
      )

    net_units_sold = _as_int(row.get("Net Units Sold"))
    avg_offer_price_usd = _convert_amount_to_usd(
      _as_float(row.get("Avg. Offer Price without tax")),
      currency_code=currency_code,
    )

    royalty_usd = _convert_amount_to_usd(
      _as_float(row.get("Royalty")),
      currency_code=currency_code,
    )
    print_cost_per_unit_usd = _convert_amount_to_usd(
      _as_float(row.get("Avg. Delivery/Manufacturing cost")),
      currency_code=currency_code,
    )

    sales_amount_usd = avg_offer_price_usd * net_units_sold
    print_cost_usd = print_cost_per_unit_usd * net_units_sold

    daily_stat = _get_or_create_daily_stat(stats_by_date, date_value)
    if format_name == "ebook":
      daily_stat.ebook_units_sold += net_units_sold
      daily_stat.ebook_royalties_usd += royalty_usd
    elif format_name == "paperback":
      daily_stat.paperback_units_sold += net_units_sold
      daily_stat.paperback_royalties_usd += royalty_usd
    else:
      daily_stat.hardcover_units_sold += net_units_sold
      daily_stat.hardcover_royalties_usd += royalty_usd

    daily_stat.total_royalties_usd += royalty_usd
    daily_stat.total_print_cost_usd += print_cost_usd

    canonical_asin = book_variant.asin
    _accumulate_sale_item(
      daily_stat.sale_items_by_asin,
      canonical_asin=canonical_asin,
      units_sold=net_units_sold,
      sales_amount_usd=sales_amount_usd,
      royalty_usd=royalty_usd,
      print_cost_usd=print_cost_usd,
    )

    market_currency_key = _market_currency_key(
      market=market,
      currency_code=currency_code,
    )
    market_currency_stats = _get_or_create_market_currency_stats(
      daily_stat=daily_stat,
      market_currency_key=market_currency_key,
      market=market,
      currency_code=currency_code,
    )
    market_currency_stats.total_units_sold += net_units_sold
    market_currency_stats.total_royalties_usd += royalty_usd
    market_currency_stats.total_print_cost_usd += print_cost_usd
    _accumulate_sale_item(
      market_currency_stats.sale_items_by_asin,
      canonical_asin=canonical_asin,
      units_sold=net_units_sold,
      sales_amount_usd=sales_amount_usd,
      royalty_usd=royalty_usd,
      print_cost_usd=print_cost_usd,
    )
    if net_units_sold > 0 and avg_offer_price_usd > 0:
      _append_price_candidate(
        market_currency_stats.avg_offer_price_usd_candidates_by_asin,
        canonical_asin=canonical_asin,
        price_usd=avg_offer_price_usd,
      )


def _apply_kenp_rows(
  rows: list[dict[str, Any]],
  stats_by_date: dict[datetime.date, models.AmazonKdpDailyStats],
) -> None:
  """Merge KENP Read rows into daily aggregates."""
  for row in rows:
    date_value = _parse_iso_date(row.get("Date"), "Date")
    asin = _required_str(row.get("ASIN"), "ASIN")
    market = _parse_market(row.get("Marketplace"))
    book_variant = book_defs.find_book_variant(asin)
    if book_variant is None:
      raise AmazonKdpError(f"Unknown ASIN in KENP Read sheet: {asin}")
    kenp_pages_read = _as_int(
      row.get("Kindle Edition Normalized Page (KENP) Read"))
    daily_stat = _get_or_create_daily_stat(stats_by_date, date_value)
    daily_stat.kenp_pages_read += kenp_pages_read

    canonical_asin = book_variant.asin
    _accumulate_sale_item(
      daily_stat.sale_items_by_asin,
      canonical_asin=canonical_asin,
      kenp_pages_read=kenp_pages_read,
    )

    currency_code = _MARKET_CODE_TO_CURRENCY_CODE.get(market, "USD")
    market_currency_key = _market_currency_key(
      market=market,
      currency_code=currency_code,
    )
    market_currency_stats = _get_or_create_market_currency_stats(
      daily_stat=daily_stat,
      market_currency_key=market_currency_key,
      market=market,
      currency_code=currency_code,
    )
    market_currency_stats.kenp_pages_read += kenp_pages_read
    _accumulate_sale_item(
      market_currency_stats.sale_items_by_asin,
      canonical_asin=canonical_asin,
      kenp_pages_read=kenp_pages_read,
    )


def _get_or_create_daily_stat(
  stats_by_date: dict[datetime.date, models.AmazonKdpDailyStats],
  date_value: datetime.date,
) -> models.AmazonKdpDailyStats:
  """Return the mutable daily stats bucket for one report date."""
  existing = stats_by_date.get(date_value)
  if existing is not None:
    return existing
  created = models.AmazonKdpDailyStats(date=date_value)
  stats_by_date[date_value] = created
  return created


def _get_or_create_market_currency_stats(
  *,
  daily_stat: models.AmazonKdpDailyStats,
  market_currency_key: str,
  market: str,
  currency_code: str,
) -> models.AmazonKdpMarketCurrencyStats:
  """Return one mutable market/currency bucket under a daily stats object."""
  existing = daily_stat.market_currency_stats_by_key.get(market_currency_key)
  if existing is not None:
    return existing
  created = models.AmazonKdpMarketCurrencyStats(
    market=market,
    currency_code=currency_code,
  )
  daily_stat.market_currency_stats_by_key[market_currency_key] = created
  return created


def _append_price_candidate(
  candidates_by_asin: dict[str, list[float]],
  *,
  canonical_asin: str,
  price_usd: float,
) -> None:
  """Append one unique per-unit price candidate for an ASIN."""
  rounded_price = round(price_usd, 6)
  if rounded_price <= 0:
    return
  prices = candidates_by_asin.setdefault(canonical_asin, [])
  if rounded_price in prices:
    return
  prices.append(rounded_price)


def _finalize_stats(
  stats_by_date: dict[datetime.date, models.AmazonKdpDailyStats], ) -> None:
  """Finalize mutable parse accumulators into deterministic persisted values."""
  for daily_stat in stats_by_date.values():
    daily_stat.total_units_sold = (daily_stat.ebook_units_sold +
                                   daily_stat.paperback_units_sold +
                                   daily_stat.hardcover_units_sold)
    daily_stat.total_royalties_usd = round(daily_stat.total_royalties_usd, 2)
    daily_stat.ebook_royalties_usd = round(daily_stat.ebook_royalties_usd, 2)
    daily_stat.paperback_royalties_usd = round(
      daily_stat.paperback_royalties_usd, 2)
    daily_stat.hardcover_royalties_usd = round(
      daily_stat.hardcover_royalties_usd, 2)
    daily_stat.total_print_cost_usd = round(daily_stat.total_print_cost_usd, 2)
    daily_stat.sale_items_by_asin = {
      asin: daily_stat.sale_items_by_asin[asin]
      for asin in sorted(daily_stat.sale_items_by_asin.keys())
    }

    daily_stat.market_currency_stats_by_key = {
      key: daily_stat.market_currency_stats_by_key[key]
      for key in sorted(daily_stat.market_currency_stats_by_key.keys())
    }
    for market_stats in daily_stat.market_currency_stats_by_key.values():
      market_stats.total_royalties_usd = round(
        market_stats.total_royalties_usd, 2)
      market_stats.total_print_cost_usd = round(
        market_stats.total_print_cost_usd, 2)
      market_stats.sale_items_by_asin = {
        asin: market_stats.sale_items_by_asin[asin]
        for asin in sorted(market_stats.sale_items_by_asin.keys())
      }
      market_stats.avg_offer_price_usd_candidates_by_asin = {
        asin: sorted(set(prices))
        for asin, prices in sorted(
          market_stats.avg_offer_price_usd_candidates_by_asin.items())
        if prices
      }


def _format_from_transaction_type(value: Any) -> str:
  """Normalize KDP transaction type to format names used in book_defs."""
  transaction_type = _required_str(value, "Transaction Type").lower()
  if "hardcover" in transaction_type:
    return "hardcover"
  if "paperback" in transaction_type:
    return "paperback"
  return "ebook"


def _parse_market(value: Any) -> str:
  """Normalize Marketplace text to a compact market code (for example `US`)."""
  marketplace = _required_str(value, "Marketplace").lower()
  market = _MARKETPLACE_TO_MARKET_CODE.get(marketplace)
  if market:
    return market
  raise AmazonKdpError(f"Unsupported Marketplace in KDP report: {value}")


def _market_currency_key(*, market: str, currency_code: str) -> str:
  """Build deterministic key for one market/currency bucket."""
  return f"{market.upper()}_{currency_code.upper()}"


def _accumulate_sale_item(
  sale_items_by_asin: dict[str, models.AmazonProductStats],
  *,
  canonical_asin: str,
  units_sold: int = 0,
  kenp_pages_read: int = 0,
  sales_amount_usd: float = 0.0,
  royalty_usd: float = 0.0,
  print_cost_usd: float = 0.0,
) -> None:
  """Accumulate one row's metrics into the target ASIN bucket."""
  existing = sale_items_by_asin.setdefault(
    canonical_asin,
    models.AmazonProductStats(
      asin=canonical_asin,
      total_print_cost_usd=0.0,
      total_royalty_usd=0.0,
    ))
  existing.units_sold += units_sold
  existing.kenp_pages_read += kenp_pages_read
  existing.total_sales_usd += sales_amount_usd
  existing.total_profit_usd += royalty_usd
  existing.total_royalty_usd = (existing.total_royalty_usd
                                or 0.0) + royalty_usd
  existing.total_print_cost_usd = (existing.total_print_cost_usd
                                   or 0.0) + print_cost_usd


def _convert_amount_to_usd(amount: float, *, currency_code: str) -> float:
  """Convert a currency amount into USD."""
  if amount == 0.0:
    return 0.0
  rate = _CURRENCY_CODE_TO_USD_RATE.get(currency_code.upper())
  if rate is None:
    raise AmazonKdpError(
      f"Unsupported currency for KDP report: {currency_code}")
  return amount * rate


def _required_str(value: Any, field_name: str) -> str:
  """Return stripped string or raise when missing."""
  result = str(value).strip() if value is not None else ""
  if not result:
    raise AmazonKdpError(f"{field_name} is required")
  return result


def _as_float(value: Any) -> float:
  """Coerce an incoming value to float."""
  if value is None:
    return 0.0
  if isinstance(value, str) and value.strip().upper() in {"", "N/A"}:
    return 0.0
  return float(value)


def _as_int(value: Any) -> int:
  """Coerce an incoming value to int."""
  return int(round(_as_float(value)))


def _parse_iso_date(value: Any, field_name: str) -> datetime.date:
  """Parse strict ISO date in yyyy-mm-dd format."""
  raw_value = _required_str(value, field_name)
  try:
    return datetime.date.fromisoformat(raw_value)
  except ValueError as exc:
    raise AmazonKdpError(
      f"{field_name} must be in YYYY-MM-DD format: {raw_value}") from exc
