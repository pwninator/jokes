"""Tests for KDP xlsx parsing helpers."""

from __future__ import annotations

import datetime
import io

import openpyxl
import pytest
from services import amazon_kdp


def _build_kdp_workbook_bytes(
  *,
  combined_sales_rows: list[list[object]],
  kenp_rows: list[list[object]],
) -> bytes:
  workbook = openpyxl.Workbook()
  default_sheet = workbook.active
  workbook.remove(default_sheet)

  combined_sheet = workbook.create_sheet("Combined Sales")
  combined_sheet.append([
    "Royalty Date",
    "Title",
    "Author Name",
    "ASIN/ISBN",
    "Marketplace",
    "Royalty Type",
    "Transaction Type",
    "Units Sold",
    "Units Refunded",
    "Net Units Sold",
    "Avg. List Price without tax",
    "Avg. Offer Price without tax",
    "Avg. Delivery/Manufacturing cost",
    "Royalty",
    "Currency",
  ])
  for row in combined_sales_rows:
    combined_sheet.append(row)

  kenp_sheet = workbook.create_sheet("KENP Read")
  kenp_sheet.append([
    "Date",
    "Title",
    "Author Name",
    "ASIN",
    "Marketplace",
    "Kindle Edition Normalized Page (KENP) Read",
  ])
  for row in kenp_rows:
    kenp_sheet.append(row)

  output = io.BytesIO()
  workbook.save(output)
  return output.getvalue()


def test_parse_kdp_xlsx_aggregates_rows_and_converts_currency():
  workbook_bytes = _build_kdp_workbook_bytes(
    combined_sales_rows=[
      [
        "2026-02-25",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0G9765J19",
        "Amazon.com",
        "35%",
        "Standard",
        1,
        0,
        1,
        2.99,
        2.99,
        "N/A",
        1.05,
        "USD",
      ],
      [
        "2026-02-25",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0GNHFKQ8W",
        "Amazon.com",
        "60%",
        "Standard - Paperback",
        1,
        0,
        1,
        11.99,
        11.99,
        2.91,
        4.28,
        "USD",
      ],
      [
        "2026-02-24",
        "Cute & Silly Valentine's Day Jokes",
        "Amelia Blanc",
        "B0GKYSMX7P",
        "Amazon.ca",
        "60%",
        "Standard - Paperback",
        2,
        0,
        2,
        16.33,
        16.33,
        8.03,
        3.54,
        "CAD",
      ],
    ],
    kenp_rows=[
      [
        "2026-02-25",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0G9765J19",
        "Amazon.com",
        55,
      ],
      [
        "2026-02-25",
        "Cute & Silly Valentine's Day Jokes",
        "Amelia Blanc",
        "B0GNMFVYC5",
        "Amazon.com",
        40,
      ],
    ],
  )

  stats = amazon_kdp.parse_kdp_xlsx(workbook_bytes)

  assert [s.date for s in stats] == [
    datetime.date(2026, 2, 24),
    datetime.date(2026, 2, 25),
  ]
  day_0224 = stats[0]
  day_0225 = stats[1]

  assert day_0224.paperback_units_sold == 2
  assert day_0224.total_units_sold == 2
  assert day_0224.total_royalties_usd == pytest.approx(2.59, abs=0.01)
  assert day_0224.total_print_cost_usd == pytest.approx(11.76, abs=0.01)

  assert day_0225.kenp_pages_read == 95
  assert day_0225.ebook_units_sold == 1
  assert day_0225.paperback_units_sold == 1
  assert day_0225.total_units_sold == 2
  assert day_0225.total_royalties_usd == pytest.approx(5.33, abs=0.01)
  assert day_0225.total_print_cost_usd == pytest.approx(2.91, abs=0.01)
  assert list(day_0225.sale_items_by_asin.keys()) == [
    "B0G9765J19",
    "B0GNHFKQ8W",
    "B0GNMFVYC5",
  ]
  ebook_item = day_0225.sale_items_by_asin["B0G9765J19"]
  assert ebook_item.kenp_pages_read == 55
  assert ebook_item.total_print_cost_usd == 0.0
  assert ebook_item.total_royalty_usd == pytest.approx(1.05, abs=0.01)
  pb_item = day_0225.sale_items_by_asin["B0GNHFKQ8W"]
  assert pb_item.total_print_cost_usd == pytest.approx(2.91, abs=0.01)
  assert pb_item.total_royalty_usd == pytest.approx(4.28, abs=0.01)
  valentines_item = day_0225.sale_items_by_asin["B0GNMFVYC5"]
  assert valentines_item.units_sold == 0
  assert valentines_item.kenp_pages_read == 40


def test_parse_kdp_xlsx_normalizes_paperback_isbn_to_variant_asin():
  workbook_bytes = _build_kdp_workbook_bytes(
    combined_sales_rows=[[
      "2026-02-25",
      "Cute & Silly Animal Jokes",
      "Amelia Blanc",
      "9798247846802",
      "Amazon.com",
      "60%",
      "Standard - Paperback",
      1,
      0,
      1,
      11.99,
      11.99,
      2.91,
      4.28,
      "USD",
    ]],
    kenp_rows=[],
  )

  stats = amazon_kdp.parse_kdp_xlsx(workbook_bytes)

  assert len(stats) == 1
  assert list(stats[0].sale_items_by_asin.keys()) == ["B0GNHFKQ8W"]


def test_parse_kdp_xlsx_persists_country_buckets_and_price_candidates():
  workbook_bytes = _build_kdp_workbook_bytes(
    combined_sales_rows=[
      [
        "2026-02-20",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0G9765J19",
        "Amazon.com",
        "35%",
        "Standard",
        1,
        0,
        1,
        2.99,
        2.99,
        "N/A",
        1.05,
        "USD",
      ],
      [
        "2026-02-20",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0GNHFKQ8W",
        "Amazon.ca",
        "60%",
        "Standard - Paperback",
        1,
        0,
        1,
        16.33,
        16.33,
        8.03,
        1.77,
        "CAD",
      ],
    ],
    kenp_rows=[
      [
        "2026-02-20",
        "Cute & Silly Animal Jokes",
        "Amelia Blanc",
        "B0G9765J19",
        "Amazon.com",
        40,
      ],
    ],
  )

  stats = amazon_kdp.parse_kdp_xlsx(workbook_bytes)

  assert len(stats) == 1
  day = stats[0]
  assert set(day.country_stats_by_code.keys()) == {"CA", "US"}

  us_bucket = day.country_stats_by_code["US"]
  assert us_bucket.country_code == "US"
  assert us_bucket.kenp_pages_read == 40
  assert us_bucket.avg_offer_price_usd_candidates_by_asin["B0G9765J19"] == [
    pytest.approx(2.99, rel=1e-6)
  ]

  ca_bucket = day.country_stats_by_code["CA"]
  assert ca_bucket.country_code == "CA"
  assert ca_bucket.total_units_sold == 1
  assert ca_bucket.avg_offer_price_usd_candidates_by_asin["B0GNHFKQ8W"] == [
    pytest.approx(11.95356, rel=1e-6)
  ]


def test_parse_kdp_xlsx_raises_on_format_mismatch():
  workbook_bytes = _build_kdp_workbook_bytes(
    combined_sales_rows=[[
      "2026-02-25",
      "Cute & Silly Animal Jokes",
      "Amelia Blanc",
      "B0G9765J19",
      "Amazon.com",
      "35%",
      "Standard - Paperback",
      1,
      0,
      1,
      2.99,
      2.99,
      0.0,
      1.05,
      "USD",
    ]],
    kenp_rows=[],
  )

  with pytest.raises(amazon_kdp.AmazonKdpError, match="Format mismatch"):
    amazon_kdp.parse_kdp_xlsx(workbook_bytes)
